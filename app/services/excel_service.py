"""
Excel Service for reading and writing Excel files.
Uses openpyxl for Excel operations.
"""

import os
import shutil
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import uuid

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)


class ExcelError(Exception):
    """Custom exception for Excel-related errors."""
    pass


class ExcelService:
    """
    Service for Excel file operations.
    Handles reading headers, writing data, and file management.
    """

    def __init__(self):
        """Initialize the Excel service."""
        self.storage_path = Path(settings.excel_storage_path)
        self.storage_path.mkdir(parents=True, exist_ok=True)

    def save_uploaded_file(
        self,
        file_content: bytes,
        original_filename: str,
        session_id: str
    ) -> Tuple[str, List[str], int]:
        """
        Save an uploaded Excel file and extract headers.

        Args:
            file_content: Raw file bytes
            original_filename: Original filename
            session_id: Session ID for organizing files

        Returns:
            Tuple of (stored_path, headers, total_rows)

        Raises:
            ExcelError: If file processing fails
        """
        try:
            # Create session directory
            session_dir = self.storage_path / session_id
            session_dir.mkdir(parents=True, exist_ok=True)

            # Generate unique filename
            ext = Path(original_filename).suffix or ".xlsx"
            stored_filename = f"data_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            stored_path = session_dir / stored_filename

            # Write file
            with open(stored_path, "wb") as f:
                f.write(file_content)

            # Read headers and count rows
            headers, total_rows = self._read_file_info(str(stored_path))

            logger.info(f"Excel file saved: {stored_path}, {len(headers)} columns, {total_rows} rows")

            return str(stored_path), headers, total_rows

        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise ExcelError(f"Failed to save Excel file: {e}")

    def _read_file_info(self, file_path: str) -> Tuple[List[str], int]:
        """
        Read headers and row count from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (headers, total_rows)
        """
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            # Read headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    break

            # Count rows (excluding header)
            total_rows = ws.max_row - 1 if ws.max_row > 0 else 0

            wb.close()

            return headers, max(0, total_rows)

        except Exception as e:
            logger.error(f"Failed to read Excel file info: {e}")
            raise ExcelError(f"Failed to read Excel file: {e}")

    def get_headers(self, file_path: str) -> List[str]:
        """
        Get column headers from an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of header names
        """
        headers, _ = self._read_file_info(file_path)
        return headers

    def write_row(
        self,
        file_path: str,
        row_number: int,
        data: Dict[str, Any],
        headers: List[str]
    ) -> bool:
        """
        Write data to a specific row in the Excel file.

        Args:
            file_path: Path to Excel file
            row_number: Row number to write (1-based, after header)
            data: Dictionary of column_name -> value
            headers: List of column headers for ordering

        Returns:
            True if successful

        Raises:
            ExcelError: If writing fails
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # Excel row number (add 1 for header row)
            excel_row = row_number + 1

            # Write each value to the correct column
            for col_idx, header in enumerate(headers, start=1):
                value = data.get(header)
                cell = ws.cell(row=excel_row, column=col_idx)

                if value is not None:
                    # Keep as string to preserve leading zeros
                    cell.value = str(value) if value else ""
                else:
                    cell.value = ""

            wb.save(file_path)
            wb.close()

            logger.info(f"Row {row_number} written to Excel")
            return True

        except Exception as e:
            logger.error(f"Failed to write row to Excel: {e}")
            raise ExcelError(f"Failed to write row: {e}")

    def read_row(
        self,
        file_path: str,
        row_number: int,
        headers: List[str]
    ) -> Dict[str, Any]:
        """
        Read a specific row from the Excel file.

        Args:
            file_path: Path to Excel file
            row_number: Row number to read (1-based, after header)
            headers: List of column headers

        Returns:
            Dictionary of column_name -> value
        """
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            excel_row = row_number + 1
            data = {}

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=excel_row, column=col_idx)
                data[header] = cell.value

            wb.close()
            return data

        except Exception as e:
            logger.error(f"Failed to read row from Excel: {e}")
            raise ExcelError(f"Failed to read row: {e}")

    def get_all_data(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        Read all data from an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (headers, list of row dictionaries)
        """
        try:
            df = pd.read_excel(file_path)
            headers = list(df.columns)
            rows = df.to_dict(orient="records")

            return headers, rows

        except Exception as e:
            logger.error(f"Failed to read all data: {e}")
            raise ExcelError(f"Failed to read Excel data: {e}")

    def create_backup(self, file_path: str) -> str:
        """
        Create a backup of an Excel file.

        Args:
            file_path: Path to original file

        Returns:
            Path to backup file
        """
        try:
            path = Path(file_path)
            backup_name = f"{path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
            backup_path = path.parent / backup_name

            shutil.copy2(file_path, backup_path)
            logger.info(f"Backup created: {backup_path}")

            return str(backup_path)

        except Exception as e:
            logger.error(f"Failed to create backup: {e}")
            raise ExcelError(f"Failed to create backup: {e}")

    def delete_session_files(self, session_id: str) -> bool:
        """
        Delete all files associated with a session.

        Args:
            session_id: Session ID

        Returns:
            True if successful
        """
        try:
            session_dir = self.storage_path / session_id
            if session_dir.exists():
                shutil.rmtree(session_dir)
                logger.info(f"Deleted session files: {session_id}")
            return True

        except Exception as e:
            logger.error(f"Failed to delete session files: {e}")
            return False

    def export_to_new_file(
        self,
        file_path: str,
        output_path: str = None
    ) -> str:
        """
        Export data to a new Excel file.

        Args:
            file_path: Source file path
            output_path: Optional output path

        Returns:
            Path to exported file
        """
        try:
            if output_path is None:
                path = Path(file_path)
                output_path = str(
                    path.parent / f"{path.stem}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
                )

            shutil.copy2(file_path, output_path)
            logger.info(f"Exported to: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Failed to export file: {e}")
            raise ExcelError(f"Failed to export: {e}")


# Global service instance
_excel_service: Optional[ExcelService] = None


def get_excel_service() -> ExcelService:
    """Get or create the Excel service instance."""
    global _excel_service
    if _excel_service is None:
        _excel_service = ExcelService()
    return _excel_service
